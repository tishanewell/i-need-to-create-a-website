from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

import openpyxl

KEY_QTY = "Consensus Dmd Plan Qty (with SOH)"
KEY_REV = "Consensus Dmd  Plan Rev (with SOH)"


def safe_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def build_records(input_path: Path) -> dict[str, dict[tuple[str, str, str, str, str], dict[str, float | str]]]:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    if "customer" not in wb.sheetnames:
        raise ValueError("Input workbook must include a 'customer' sheet.")

    ws = wb["customer"]
    records: dict[str, dict[tuple[str, str, str, str, str], dict[str, float | str]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 21:
            continue
        key_figure = row[0]
        brand = row[1]
        product_line = row[2]
        product_id = row[3]
        product_desc = row[4]
        customer = row[5]
        product_status = row[6]
        total_2026 = safe_number(row[20])

        if not customer or not product_id:
            continue
        if key_figure not in (KEY_QTY, KEY_REV):
            continue

        customer_map = records.setdefault(customer, {})
        product_key = (
            str(brand or ""),
            str(product_line or ""),
            str(product_id or ""),
            str(product_desc or ""),
            str(product_status or ""),
        )
        product_record = customer_map.setdefault(product_key, {"qty": 0.0, "rev": 0.0})
        if key_figure == KEY_QTY:
            product_record["qty"] = product_record["qty"] + total_2026
        else:
            product_record["rev"] = product_record["rev"] + total_2026
    return records


def write_customer_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    customer_records: dict[tuple[str, str, str, str, str], dict[str, float | str]],
) -> None:
    style_row = 3 if ws.max_row >= 3 else 2
    base_styles = [copy(ws.cell(style_row, c)._style) for c in range(1, 10)]
    base_formats = [ws.cell(style_row, c).number_format for c in range(1, 10)]
    base_alignment = [copy(ws.cell(style_row, c).alignment) for c in range(1, 10)]

    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    sorted_items = sorted(customer_records.items(), key=lambda x: (x[0][0], x[0][1], x[0][2]))
    for idx, (product_key, values) in enumerate(sorted_items, start=3):
        brand, product_line, product_id, product_desc, product_status = product_key
        qty = values.get("qty", 0.0)
        rev = values.get("rev", 0.0)
        row_values = [
            brand,
            product_line,
            product_id,
            product_desc,
            product_status,
            "[TBD]",
            "[TBD]",
            qty if qty else None,
            rev if rev else None,
        ]
        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(idx, col, value=value)
            cell._style = copy(base_styles[col - 1])
            cell.number_format = base_formats[col - 1]
            cell.alignment = copy(base_alignment[col - 1])


def convert_file(input_path: Path, template_path: Path, output_path: Path) -> dict[str, Any]:
    if not template_path.exists():
        raise FileNotFoundError(f"Template file is missing: {template_path}")

    records = build_records(input_path)
    wb = openpyxl.load_workbook(template_path)

    sheet_count = 0
    row_count = 0
    for ws in wb.worksheets:
        customer_name = ws.title
        customer_records = records.get(customer_name, {})
        write_customer_sheet(ws, customer_records)
        sheet_count += 1
        row_count += len(customer_records)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {"sheets_updated": sheet_count, "rows_written": row_count}
